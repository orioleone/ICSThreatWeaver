from pathlib import Path

from openpyxl import Workbook, load_workbook

from backend.app.workbook_transformer import transform_risk_assessment_workbook


def _build_input_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    # Row 3: headers (matching the updated ref template layout)
    ws["A3"] = "Risk ID"
    ws["B3"] = "Zone"
    ws["C3"] = "Asset"
    ws["D3"] = "Threat Source"
    ws["E3"] = "Tactics"
    ws["F3"] = "ID"
    ws["G3"] = "Technique"
    ws["H3"] = "Description"
    ws["I3"] = "Vulnerabilities"
    ws["J3"] = "Consequence Description"
    ws["K3"] = "S"
    ws["L3"] = "E"
    ws["M3"] = "F"
    ws["N3"] = "R"
    ws["O3"] = "Max"
    ws["S3"] = "Countermeasures"
    # Row 4: data
    ws["A4"] = 101
    ws["B4"] = "Control Zone"
    ws["C4"] = "Control Server"
    ws["D4"] = "External - Authorised - Vendor"
    ws["E4"] = "Initial Access"
    ws["F4"] = "T0822"
    ws["G4"] = "External Remote Services"
    ws["H4"] = "Remote access can be used for initial access."
    ws["I4"] = "Unpatched remote access gateway"
    ws["J4"] = "Loss of visibility"
    ws["K4"] = 4
    ws["L4"] = 2
    ws["M4"] = 1
    ws["N4"] = 3
    ws["O4"] = 4
    ws["S4"] = "Restrict remote access; enforce MFA"
    wb.save(path)
    wb.close()


def _build_template_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Risk Assessment"
    ws["A1"] = "Template"
    wb.create_sheet("Mapping")
    wb.save(path)
    wb.close()


def _build_builder_workbook(path: Path):
    wb = Workbook()
    config = wb.active
    config.title = "Config"
    config["A2"] = "ID"
    config["B2"] = "Name"
    config["A3"] = "A0007"
    config["B3"] = "Control Server"
    matrix = wb.create_sheet("ICSMatrix")
    matrix.append(["Initial Access", "T0822", "External Remote Services", "Vendor Remote Access", "desc", "A0007"])
    wb.save(path)
    wb.close()


def test_transform_workbook_creates_required_sheets(tmp_path):
    source = tmp_path / "input.xlsx"
    template = tmp_path / "template.xlsm"
    builder = tmp_path / "builder.xlsm"
    output = tmp_path / "out.xlsm"

    _build_input_workbook(source)
    _build_template_workbook(template)
    _build_builder_workbook(builder)

    result = transform_risk_assessment_workbook(
        source_workbook_path=str(source),
        template_workbook_path=str(template),
        mitre_builder_workbook_path=str(builder),
        output_path=str(output),
    )

    assert Path(result["output_path"]).exists()
    assert result["enhanced_rows"] >= 1

    wb = load_workbook(output, keep_vba=True)
    try:
        assert "Enhanced Risk Assessment" in wb.sheetnames
        assert "Mapping Reference" in wb.sheetnames
        assert "MITRE Reference Index" in wb.sheetnames
        assert "Change Log" in wb.sheetnames
        assert wb["Enhanced Risk Assessment"]["A2"].value == 1  # sequential risk_id
    finally:
        wb.close()
