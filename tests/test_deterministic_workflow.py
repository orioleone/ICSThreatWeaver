from pathlib import Path

from openpyxl import Workbook, load_workbook

from backend.app.excel_export import export_matrix_to_workbook
from backend.app.mapping_engine import MappingEngine


def test_mapping_engine_builds_deterministic_technique_to_asset_report():
    engine = MappingEngine()
    techniques = [
        {"external_id": "T0800", "name": "Activate Firmware Update Mode", "description": "controller and hmi interaction"},
        {"external_id": "T0801", "name": "Monitor Process State", "description": "historian and network server telemetry"},
    ]

    report = engine.build_deterministic_mapping_report(techniques)

    assert report["coverage_summary"]["technique_count"] == 2
    assert report["mappings"][0]["technique_id"] == "T0800"
    assert "Human-Machine Interface (HMI)" in report["mappings"][0]["mapped_assets"]


def test_zone_matrix_aggregates_assets_per_technique():
    engine = MappingEngine()
    matrix = engine.generate_zone_matrix(
        zone_name="Zone A",
        selected_assets=["Control Server", "Human-Machine Interface (HMI)"],
        approved_asset_technique_map={
            "Control Server": [
                {"external_id": "T0822", "name": "External Remote Services", "description": "desc", "tactics": ["Initial Access"], "mitigations": ["Network Segmentation"]}
            ],
            "Human-Machine Interface (HMI)": [
                {"external_id": "T0822", "name": "External Remote Services", "description": "desc", "tactics": ["Initial Access"], "mitigations": ["Network Segmentation", "User Account Management"]}
            ],
        },
    )

    assert len(matrix) == 1
    assert matrix[0]["zone"] == "Zone A"
    assert matrix[0]["asset"] == "Control Server; Human-Machine Interface (HMI)"


def test_excel_export_uses_single_blank_template_schema(tmp_path):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws["A3"] = "Zone"
    ws["B3"] = "Asset"
    ws["C3"] = "Technique ID"
    ws["D3"] = "Technique Name"
    ws["E3"] = "Description"
    ws["F3"] = "Mitigations"
    wb.save(template)

    export_matrix_to_workbook(
        matrix_rows=[{
            "zone": "Zone A",
            "asset": "Control Server",
            "technique_id": "T0822",
            "technique_name": "External Remote Services",
            "description": "desc",
            "mitigations": "Network Segmentation",
        }],
        template_path=str(template),
        output_path=str(output),
        sheet_name="Worksheet",
        start_cell="A4",
    )

    out_wb = load_workbook(output)
    try:
        out_ws = out_wb["Worksheet"]
        assert out_ws["A4"].value == "Zone A"
        assert out_ws["B4"].value == "Control Server"
        assert out_ws["C4"].value == "T0822"
    finally:
        out_wb.close()
