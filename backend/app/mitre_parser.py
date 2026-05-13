from __future__ import annotations

import logging
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from sqlalchemy import desc
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

from .models import (
    DatasetVersion,
    MitreAsset,
    MitreMitigation,
    MitreTactic,
    MitreTechnique,
    TechniqueMapping,
    TechniqueMitigation,
    TechniqueTactic,
    DEFAULT_MAPPING_VERSION,
)

DEFAULT_MITRE_ICS_URL = "https://raw.githubusercontent.com/mitre-attack/attack-stix-data/master/ics-attack/ics-attack.json"
BASE_DIR = Path(__file__).resolve().parents[2]
DEFAULT_MITRE_ASSET_BUILDER = BASE_DIR / "docs" / "MITRE ATTACK Techniques Builder.xlsm"

DEFAULT_ICS_ASSET_CATALOG = [
    {"external_id": "A0008", "name": "Application Server", "description": "Application-layer servers supporting industrial operations."},
    {"external_id": "A0007", "name": "Control Server", "description": "Supervisory control servers coordinating industrial processes."},
    {"external_id": "A0009", "name": "Data Gateway", "description": "Systems that broker or relay industrial data between network segments."},
    {"external_id": "A0006", "name": "Data Historian", "description": "Repositories for operational and time-series process data."},
    {"external_id": "A0013", "name": "Field I/O", "description": "Input and output devices connected to controllers and field equipment."},
    {"external_id": "A0002", "name": "Human-Machine Interface (HMI)", "description": "Operator interface systems used to monitor and control the process."},
    {"external_id": "A0005", "name": "Intelligent Electronic Device (IED)", "description": "Protection or automation devices commonly used in power and industrial systems."},
    {"external_id": "A0012", "name": "Jump Host", "description": "Brokered access systems used to administer or reach secured industrial assets."},
    {"external_id": "A0003", "name": "Programmable Logic Controller (PLC)", "description": "Industrial control devices executing logic and interfacing with field equipment."},
    {"external_id": "A0004", "name": "Remote Terminal Unit (RTU)", "description": "Remote monitoring and control units used in distributed industrial systems."},
    {"external_id": "A0014", "name": "Routers", "description": "Network routing infrastructure connecting industrial segments and conduits."},
    {"external_id": "A0017", "name": "Distributed Control System (DCS) Controller", "description": "Distributed control controllers used to coordinate industrial process logic."},
    {"external_id": "A0018", "name": "Programmable Automation Controller (PAC)", "description": "Industrial automation controllers used for advanced modular process control."},
    {"external_id": "A0010", "name": "Safety Controller", "description": "Safety instrumented controllers enforcing protective shutdown logic."},
    {"external_id": "A0011", "name": "Virtual Private Network (VPN) Server", "description": "Remote access gateways supporting authenticated industrial connectivity."},
    {"external_id": "A0015", "name": "Switch", "description": "Network switches providing layer-2 connectivity within industrial network segments."},
    {"external_id": "A0016", "name": "Firewall", "description": "Network firewalls enforcing traffic policies and segmentation between industrial zones and conduits."},
    {"external_id": "A0001", "name": "Workstation", "description": "Engineering or operator workstations used to manage industrial assets."},
]


def _extract_external_id(obj: dict[str, Any]) -> str | None:
    for ref in obj.get("external_references", []):
        if ref.get("external_id"):
            return ref["external_id"]
    return None


def _is_live_attack_object(obj: dict[str, Any]) -> bool:
    return not obj.get("revoked", False) and not obj.get("x_mitre_deprecated", False)


def download_mitre_ics_bundle(url: str = DEFAULT_MITRE_ICS_URL) -> dict[str, Any]:
    """Fetch the MITRE ATT&CK for ICS STIX bundle from *url*.

    Raises ``requests.HTTPError`` on a non-2xx response and
    ``requests.Timeout`` if the server takes longer than 60 seconds.
    """
    logger.info("Fetching MITRE ICS bundle from %s", url)
    # [RESOLVED] allow_redirects=False is an intentional SSRF safeguard: we only
    # trust the explicitly configured upstream host; following redirects could send
    # the request to an attacker-controlled server.
    # [SOURCE] Audit finding: Low (L-5)
    attempts = 3
    transient_statuses = {429, 500, 502, 503, 504}
    for attempt in range(1, attempts + 1):
        try:
            response = requests.get(url, timeout=60, allow_redirects=False)
            if response.status_code in transient_statuses:
                if attempt >= attempts:
                    response.raise_for_status()
                backoff_seconds = attempt
                logger.warning(
                    "Transient MITRE fetch HTTP %d (attempt %d/%d). Retrying in %ds",
                    response.status_code,
                    attempt,
                    attempts,
                    backoff_seconds,
                )
                time.sleep(backoff_seconds)
                continue
            response.raise_for_status()
            return response.json()
        except (requests.Timeout, requests.ConnectionError) as exc:
            if attempt >= attempts:
                logger.error("Failed to fetch MITRE ICS bundle after %d attempts", attempts, exc_info=True)
                raise
            backoff_seconds = attempt
            logger.warning(
                "Transient MITRE fetch failure (attempt %d/%d): %s. Retrying in %ds",
                attempt,
                attempts,
                exc,
                backoff_seconds,
            )
            time.sleep(backoff_seconds)


def parse_stix_bundle(bundle: dict[str, Any]) -> dict[str, Any]:
    objects = bundle.get("objects", [])
    by_stix_id = {obj.get("id"): obj for obj in objects if obj.get("id")}

    parsed: dict[str, Any] = {
        "techniques": [],
        "tactics": [],
        "mitigations": [],
        "technique_tactics": [],
        "technique_mitigations": [],
        "stix_assets": [],
        "technique_asset_targets": [],
    }

    for obj in objects:
        if not _is_live_attack_object(obj):
            continue

        obj_type = obj.get("type")
        if obj_type == "x-mitre-asset":
            external_id = _extract_external_id(obj)
            if external_id:
                parsed["stix_assets"].append(
                    {
                        "stix_id": obj["id"],
                        "external_id": external_id,
                        "name": obj.get("name", ""),
                        "description": obj.get("description", ""),
                    }
                )
        elif obj_type == "x-mitre-tactic":
            parsed["tactics"].append(
                {
                    "stix_id": obj["id"],
                    "external_id": _extract_external_id(obj),
                    "name": obj.get("name", ""),
                    "shortname": obj.get("x_mitre_shortname", obj.get("name", "").lower().replace(" ", "-")),
                    "description": obj.get("description", ""),
                }
            )
        elif obj_type == "attack-pattern":
            external_id = _extract_external_id(obj)
            if not external_id:
                continue
            tactic_names = [phase.get("phase_name") for phase in obj.get("kill_chain_phases", []) if phase.get("phase_name")]
            parsed["techniques"].append(
                {
                    "stix_id": obj["id"],
                    "external_id": external_id,
                    "name": obj.get("name", ""),
                    "description": obj.get("description", ""),
                    "modified": obj.get("modified"),
                    "platforms": ", ".join(obj.get("x_mitre_platforms", [])),
                    "is_subtechnique": obj.get("x_mitre_is_subtechnique", False),
                    "tactics": tactic_names,
                }
            )
            for tactic_name in tactic_names:
                parsed["technique_tactics"].append((external_id, tactic_name))
        elif obj_type == "course-of-action":
            external_id = _extract_external_id(obj)
            parsed["mitigations"].append(
                {
                    "stix_id": obj["id"],
                    "external_id": external_id,
                    "name": obj.get("name", ""),
                    "description": obj.get("description", ""),
                }
            )

    for obj in objects:
        if obj.get("type") != "relationship":
            continue
        # Skip revoked or deprecated relationship objects
        if not _is_live_attack_object(obj):
            continue

        rel_type = obj.get("relationship_type")
        source = by_stix_id.get(obj.get("source_ref"), {})
        target = by_stix_id.get(obj.get("target_ref"), {})

        # Skip if either endpoint is missing from the bundle or is itself revoked/deprecated
        if not source or not target:
            continue
        if not _is_live_attack_object(source) or not _is_live_attack_object(target):
            continue

        if rel_type == "mitigates":
            mitigation_id = _extract_external_id(source)
            technique_id = _extract_external_id(target)
            if mitigation_id and technique_id:
                parsed["technique_mitigations"].append((technique_id, mitigation_id))

        elif rel_type == "targets":
            # attack-pattern targets x-mitre-asset
            if source.get("type") == "attack-pattern" and target.get("type") == "x-mitre-asset":
                technique_id = _extract_external_id(source)
                asset_id = _extract_external_id(target)
                if technique_id and asset_id:
                    parsed["technique_asset_targets"].append((technique_id, asset_id))

    return parsed


def diff_techniques(old: dict[str, dict[str, Any]], new: dict[str, dict[str, Any]]) -> dict[str, list[str]]:
    old_ids = set(old)
    new_ids = set(new)

    new_techniques = sorted(new_ids - old_ids)
    modified = sorted(
        technique_id
        for technique_id in new_ids & old_ids
        if any(
            old[technique_id].get(field) != new[technique_id].get(field)
            for field in ("modified", "name", "description")
        )
    )
    return {"new": new_techniques, "modified": modified}


def load_mitre_asset_catalog(builder_path: str | None = None) -> list[dict[str, str]]:
    candidate_path = Path(builder_path) if builder_path else DEFAULT_MITRE_ASSET_BUILDER
    if candidate_path.exists():
        workbook = load_workbook(candidate_path, keep_vba=candidate_path.suffix.lower() == ".xlsm", data_only=False)
        try:
            if "Config" in workbook.sheetnames:
                sheet = workbook["Config"]
                assets: list[dict[str, str]] = []
                for row in sheet.iter_rows(min_row=3, values_only=True):
                    if row and row[0] and row[1]:
                        external_id = str(row[0]).strip()
                        name = str(row[1]).strip()
                        # Skip header rows (e.g. "ID" / "Name" or "ID - Name")
                        if external_id.lower() in ("id", "id - name") or name.lower() in ("name", "id - name"):
                            continue
                        description = next(
                            (item["description"] for item in DEFAULT_ICS_ASSET_CATALOG if item["external_id"] == external_id),
                            f"MITRE ATT&CK for ICS asset: {name}",
                        )
                        assets.append({"external_id": external_id, "name": name, "description": description})
                if assets:
                    return assets
        finally:
            workbook.close()
    return list(DEFAULT_ICS_ASSET_CATALOG)


def seed_builtin_assets(db: Session) -> None:
    for asset in load_mitre_asset_catalog():
        exists = db.query(MitreAsset).filter(
            (MitreAsset.external_id == asset["external_id"]) | (MitreAsset.name == asset["name"])
        ).first()
        if not exists:
            # Bootstrap only — insert rows that do not exist yet.
            # Existing rows are left untouched so STIX import data is always authoritative.
            db.add(MitreAsset(external_id=asset["external_id"], name=asset["name"], description=asset["description"]))
    db.commit()


def store_bundle(db: Session, bundle: dict[str, Any], version_tag: str | None = None, source_url: str = DEFAULT_MITRE_ICS_URL) -> dict[str, Any]:
    parsed = parse_stix_bundle(bundle)
    seed_builtin_assets(db)

    if not version_tag:
        # [RESOLVED] Using timezone-aware datetime.now() to avoid deprecated utcnow().
        # Compatible with Python 3.12+.
        # [SOURCE] Audit finding: High (H-4)
        version_tag = datetime.now(timezone.utc).strftime("mitre-ics-%Y%m%d-%H%M%S")

    previous_version = db.query(DatasetVersion).order_by(desc(DatasetVersion.imported_at)).first()
    previous_snapshot: dict[str, dict[str, Any]] = {}
    if previous_version:
        previous_techniques = db.query(MitreTechnique).filter(MitreTechnique.dataset_version_id == previous_version.id).all()
        previous_snapshot = {
            technique.external_id: {
                "modified": technique.modified,
                "name": technique.name,
                "description": technique.description,
            }
            for technique in previous_techniques
        }

    new_snapshot = {
        technique["external_id"]: {
            "modified": technique.get("modified"),
            "name": technique.get("name"),
            "description": technique.get("description"),
        }
        for technique in parsed["techniques"]
    }
    diff = diff_techniques(previous_snapshot, new_snapshot) if previous_snapshot else {"new": sorted(new_snapshot.keys()), "modified": []}

    dataset_version = DatasetVersion(version_tag=version_tag, source_url=source_url)
    db.add(dataset_version)
    # [RESOLVED] Use flush() instead of commit() so the DatasetVersion row receives
    # a primary key within the current transaction WITHOUT committing it.  If any
    # subsequent insert fails, the caller's exception handler (or SQLAlchemy's own
    # rollback) will remove the orphaned version record, keeping the DB consistent.
    # [SOURCE] Audit finding: High (H-3)
    db.flush()
    db.refresh(dataset_version)

    tactic_lookup: dict[str, MitreTactic] = {}
    for tactic in parsed["tactics"]:
        row = MitreTactic(dataset_version_id=dataset_version.id, **tactic)
        db.add(row)
        db.flush()
        tactic_lookup[tactic["shortname"]] = row

    technique_lookup: dict[str, MitreTechnique] = {}
    for technique in parsed["techniques"]:
        row = MitreTechnique(
            dataset_version_id=dataset_version.id,
            stix_id=technique["stix_id"],
            external_id=technique["external_id"],
            name=technique["name"],
            description=technique.get("description"),
            modified=technique.get("modified"),
            platforms=technique.get("platforms"),
            is_subtechnique=technique.get("is_subtechnique", False),
            is_new=technique["external_id"] in diff["new"],
            is_modified=technique["external_id"] in diff["modified"],
        )
        db.add(row)
        db.flush()
        technique_lookup[technique["external_id"]] = row

    mitigation_lookup: dict[str, MitreMitigation] = {}
    for mitigation in parsed["mitigations"]:
        row = MitreMitigation(dataset_version_id=dataset_version.id, **mitigation)
        db.add(row)
        db.flush()
        if mitigation.get("external_id"):
            mitigation_lookup[mitigation["external_id"]] = row

    for technique_external_id, tactic_name in parsed["technique_tactics"]:
        technique = technique_lookup.get(technique_external_id)
        tactic = tactic_lookup.get(tactic_name)
        if technique and tactic:
            db.add(TechniqueTactic(technique_id=technique.id, tactic_id=tactic.id))

    for technique_external_id, mitigation_external_id in parsed["technique_mitigations"]:
        technique = technique_lookup.get(technique_external_id)
        mitigation = mitigation_lookup.get(mitigation_external_id)
        if technique and mitigation:
            db.add(TechniqueMitigation(technique_id=technique.id, mitigation_id=mitigation.id))

    db.commit()

    # Upsert STIX-defined assets (override catalog entries where external_id matches)
    for stix_asset in parsed["stix_assets"]:
        # Skip any header-like placeholder entries
        if stix_asset["name"].lower() in ("id - name", "name") or stix_asset["external_id"].lower() in ("id", "id - name"):
            continue
        existing = db.query(MitreAsset).filter(MitreAsset.external_id == stix_asset["external_id"]).first()
        if existing:
            existing.name = stix_asset["name"]
            existing.description = stix_asset["description"]
            existing.source = "MITRE"
        else:
            db.add(MitreAsset(
                external_id=stix_asset["external_id"],
                name=stix_asset["name"],
                description=stix_asset["description"],
                source="MITRE",
            ))
    db.commit()

    # Build asset lookup by external_id for STIX targets relationships
    asset_lookup: dict[str, MitreAsset] = {
        asset.external_id: asset
        for asset in db.query(MitreAsset).filter(MitreAsset.external_id.is_not(None)).all()
    }

    # Store STIX-derived technique→asset mappings (source="stix")
    stix_links_created = 0
    for technique_external_id, asset_external_id in parsed["technique_asset_targets"]:
        technique = technique_lookup.get(technique_external_id)
        asset = asset_lookup.get(asset_external_id)
        if not technique or not asset:
            continue
        existing = db.query(TechniqueMapping).filter(
            TechniqueMapping.technique_id == technique.id,
            TechniqueMapping.mitre_asset_id == asset.id,
            # [IMPROVEMENT] Use DEFAULT_MAPPING_VERSION constant instead of "v1" literal.
            # [SOURCE] Audit finding: Low (L-1)
            TechniqueMapping.mapping_version == DEFAULT_MAPPING_VERSION,
        ).first()
        if not existing:
            db.add(TechniqueMapping(
                mitre_asset_id=asset.id,
                technique_id=technique.id,
                mapping_version=DEFAULT_MAPPING_VERSION,
                source="stix",
                justification="MITRE ATT&CK STIX targets relationship",
                confidence=1.0,
                approved=True,
            ))
            stix_links_created += 1
    db.commit()

    # Invariant enforcement: after writing STIX rows, remove every rule-sourced
    # mapping for any asset that STIX covers. If STIX defines any techniques
    # for an asset its mapping is authoritative and complete; keyword rules
    # must not add to it. Applies to all assets and all historical versions —
    # no asset or technique ID is hardcoded here.
    stix_covered_asset_ids = {
        tm.mitre_asset_id
        for tm in db.query(TechniqueMapping).filter(TechniqueMapping.source == "stix").all()
    }
    if stix_covered_asset_ids:
        db.query(TechniqueMapping).filter(
            TechniqueMapping.source == "rule",
            TechniqueMapping.mitre_asset_id.in_(stix_covered_asset_ids),
        ).delete(synchronize_session=False)
        db.commit()

    asset_count = db.query(MitreAsset).filter(MitreAsset.external_id.is_not(None)).count()

    return {
        "version_tag": version_tag,
        "technique_count": len(parsed["techniques"]),
        "tactic_count": len(parsed["tactics"]),
        "mitigation_count": len(parsed["mitigations"]),
        "asset_count": asset_count,
        "stix_asset_links": stix_links_created,
        "new_techniques": diff["new"],
        "modified_techniques": diff["modified"],
    }
