from __future__ import annotations

import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .mitre_parser import download_mitre_ics_bundle, parse_stix_bundle

logger = logging.getLogger(__name__)


ASSET_ID_PATTERN = re.compile(r"A\d{4}")
TECHNIQUE_ID_PATTERN = re.compile(r"T\d{4}(?:\.\d{3})?")

# [RESOLVED] ZIP bomb defense: Validate workbook structure to prevent malicious
# archives from consuming excessive memory during decompression.
# [SOURCE] Audit finding: High (H-5)
def _validate_workbook_structure(workbook, max_sheets: int = 100, max_rows_per_sheet: int = 100000) -> None:
    """Validate that workbook structure is reasonable (not a ZIP bomb).
    
    Args:
        workbook: Loaded openpyxl Workbook instance.
        max_sheets: Maximum number of sheets allowed (default 100).
        max_rows_per_sheet: Maximum rows per sheet (default 100,000).
        
    Raises:
        ValueError: If workbook structure exceeds limits.
    """
    if len(workbook.sheetnames) > max_sheets:
        raise ValueError(
            f"Workbook contains {len(workbook.sheetnames)} sheets; maximum allowed is {max_sheets}. "
            "This may indicate a malicious or corrupted file."
        )
    
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        if ws.max_row and ws.max_row > max_rows_per_sheet:
            raise ValueError(
                f"Sheet '{sheet}' contains {ws.max_row} rows; maximum allowed is {max_rows_per_sheet}. "
                "This may indicate a malicious or corrupted file."
            )

# [RESOLVED] Formula-injection guard: same pattern as excel_export._safe_cell_value.
# [SOURCE] Audit finding: Critical (C-2)
# Defined here independently to avoid a cross-module import between transformer and exporter.
# MITRE STIX descriptions and builder-workbook content are external inputs that could
# contain strings beginning with '=', '+', '-', '@' which spreadsheet apps evaluate as formulas.
# [SOURCE] Audit finding: Critical (C-2)
_FORMULA_PREFIX_RE = re.compile(r"^[=+\-@]")


def _safe_cell_value(value: Any) -> Any:
    """Escape strings that start with a spreadsheet formula character.

    Returns the value with a leading apostrophe when it would otherwise be
    interpreted as a formula by Excel / LibreOffice.  Non-string values are
    returned unchanged.
    """
    if isinstance(value, str) and _FORMULA_PREFIX_RE.match(value):
        return "'" + value
    return value

# ---------------------------------------------------------------------------
# Column structure constants for the Enhanced Risk Assessment output sheet
# ---------------------------------------------------------------------------

# Mapping from internal field name to the 1-based column index in the source
# "Worksheet" sheet, as laid out by the ISA 62443-3-2 DRA template.
_FIELD_TO_SOURCE_COL: dict[str, int] = {
    "risk_id": 1,               # Column A
    "zone": 2,                  # Column B
    "asset": 3,                 # Column C
    "original_threat_source": 4,  # Column D
    "tactics": 5,               # Column E
    "technique_id": 6,          # Column F
    "technique_name": 7,        # Column G
    "technique_description": 8, # Column H
    "vulnerabilities": 9,       # Column I
    "consequence_description": 10,  # Column J
    "countermeasures_template": 19,  # Column S
}

# Human-readable labels for MITRE-enrichment fields added by this tool that
# are not present in the source workbook.
_ENRICHMENT_FIELD_LABELS: dict[str, str] = {
    "adversary": "Adversary",
    "adversary_type": "Adversary Type",
    "adversary_description": "Adversary Description",
    "mitigations": "MITRE Mitigations",
    "impact": "Impact",
    "source_row": "Source Row",
}

# Ordered list of field keys written to the Enhanced Risk Assessment sheet.
_ENHANCED_FIELD_ORDER: list[str] = [
    "risk_id",
    "zone",
    "asset",
    "adversary",
    "adversary_type",
    "adversary_description",
    "technique_id",
    "technique_name",
    "technique_description",
    "tactics",
    "mitigations",
    "countermeasures_template",
    "impact",
    "original_threat_source",
    "vulnerabilities",
    "consequence_description",
    "source_row",
]


def _normalize_adversary(threat_source: str | None) -> tuple[str, str, str]:
    """Classify a free-text threat-source label into a structured adversary model.

    Args:
        threat_source: Raw threat-source string from the source workbook, or
            ``None`` if the cell was empty.

    Returns:
        A three-tuple of ``(adversary_name, adversary_type, adversary_description)``.
    """
    text = (threat_source or "N/A").strip()
    lowered = text.lower()

    if "nation" in lowered or "warrior" in lowered or "state" in lowered or "apt" in lowered:
        return (text if text != "N/A" else "Nation-state", "APT", "Potential state-aligned or advanced persistent threat actor.")
    if "insider" in lowered or "internal" in lowered:
        return (text if text != "N/A" else "Insider", "Insider", "Internal personnel or trusted third-party with contextual access.")
    if "vendor" in lowered or "contractor" in lowered or "third party" in lowered:
        return (text, "Generic", "Trusted external support or vendor-origin threat source.")
    if "criminal" in lowered or "malware" in lowered or "ransom" in lowered:
        return (text if text != "N/A" else "Cybercriminal", "Criminal", "Financially motivated or opportunistic cybercriminal threat.")
    if "hacktiv" in lowered:
        return (text if text != "N/A" else "Hacktivist", "Generic", "Ideologically motivated external threat source.")
    if text == "N/A":
        return ("Generic External Threat", "Generic", "Fallback adversary classification used where direct mapping is unavailable.")
    return (text, "Generic", "Normalized from the original threat source text.")


def _extract_asset_ids(zone_text: Any) -> list[str]:
    """Extract MITRE ICS asset identifiers (e.g. A0007) from a zone cell value.

    Args:
        zone_text: Raw zone cell value which may embed one or more asset IDs.

    Returns:
        Sorted, deduplicated list of asset ID strings found in the text.
    """
    return sorted(set(ASSET_ID_PATTERN.findall(str(zone_text or ""))))


def _load_builder_catalog(builder_path: str) -> tuple[dict[str, str], dict[str, dict[str, Any]]]:
    """Load the asset and technique catalogs from the MITRE ATT&CK Builder workbook.

    The builder workbook is expected to contain two sheets:
    - ``Config``: rows from row 3 with columns [asset_id, asset_name].
    - ``ICSMatrix``: one technique per row with columns
      [tactic, technique_id, name, adversary, description, ...asset_ids].

    Args:
        builder_path: Absolute or relative path to the ``.xlsm`` builder file.

    Returns:
        A two-tuple of ``(asset_catalog, technique_catalog)`` where
        ``asset_catalog`` maps asset ID strings to display names, and
        ``technique_catalog`` maps technique IDs to dicts with keys
        ``tactic``, ``technique_name``, ``default_adversary``, ``description``,
        and ``asset_ids``.

    Raises:
        FileNotFoundError: Propagated from :func:`openpyxl.load_workbook` if
            ``builder_path`` does not exist.
    """
    wb = load_workbook(builder_path, keep_vba=True, data_only=False)
    _validate_workbook_structure(wb, max_sheets=10, max_rows_per_sheet=10000)

    asset_catalog: dict[str, str] = {}
    if "Config" in wb.sheetnames:
        ws = wb["Config"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row and row[0] and row[1]:
                asset_catalog[str(row[0]).strip()] = str(row[1]).strip()

    technique_catalog: dict[str, dict[str, Any]] = {}
    if "ICSMatrix" in wb.sheetnames:
        ws = wb["ICSMatrix"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or not row[1]:
                continue
            technique_id = str(row[1]).strip()
            if not TECHNIQUE_ID_PATTERN.fullmatch(technique_id):
                continue
            asset_ids = [str(value).strip() for value in row[5:] if value]
            technique_catalog[technique_id] = {
                "tactic": str(row[0] or ""),
                "technique_name": str(row[2] or ""),
                "default_adversary": str(row[3] or ""),
                "description": str(row[4] or ""),
                "asset_ids": asset_ids,
            }

    wb.close()

    # [IMPROVEMENT] Warn when expected sheets are absent or empty so callers can
    # detect a mis-configured builder workbook instead of silently falling back.
    # [SOURCE] Audit finding: Medium (M-6)
    if not asset_catalog:
        logger.warning(
            "Builder workbook '%s': 'Config' sheet is missing or contains no asset rows."
            " Asset ID resolution will fall back to raw IDs.",
            builder_path,
        )
    if not technique_catalog:
        logger.warning(
            "Builder workbook '%s': 'ICSMatrix' sheet is missing or contains no technique rows."
            " MITRE technique enrichment from the builder will be unavailable.",
            builder_path,
        )

    return asset_catalog, technique_catalog


def _load_mitre_reference() -> tuple[dict[str, dict[str, Any]], dict[str, list[str]]]:
    """Download and parse the live MITRE ATT&CK for ICS STIX bundle.

    Performs a best-effort fetch; if the download or parse fails for any
    reason (network unavailable, upstream format change, etc.) the function
    returns empty dicts so that the pipeline can continue with local data.

    Returns:
        A two-tuple of ``(techniques, mitigation_map)`` where ``techniques``
        maps technique IDs to canonical STIX technique dicts and
        ``mitigation_map`` maps technique IDs to lists of mitigation names.
    """
    try:
        bundle = download_mitre_ics_bundle()
        parsed = parse_stix_bundle(bundle)
    except Exception as exc:
        # [IMPROVEMENT] Log the full exception before returning empty dicts.  Silent
        # fallback previously produced degraded output with no user notification,
        # which is a compliance risk when MITRE enrichment data is required as
        # evidence for an ISA/IEC 62443-3-2 DRA.
        # [SOURCE] Audit finding: Medium (M-7)
        logger.warning(
            "MITRE ICS STIX fetch failed — proceeding without live enrichment: %s",
            exc,
            exc_info=True,
        )
        return {}, {}

    techniques = {item["external_id"]: item for item in parsed["techniques"]}
    mitigation_names = {item.get("external_id"): item.get("name", "") for item in parsed["mitigations"] if item.get("external_id")}
    mitigation_map: dict[str, list[str]] = defaultdict(list)
    for technique_id, mitigation_id in parsed["technique_mitigations"]:
        name = mitigation_names.get(mitigation_id, mitigation_id)
        if name and name not in mitigation_map[technique_id]:
            mitigation_map[technique_id].append(name)
    return techniques, dict(mitigation_map)


def _find_source_sheet(source_path: str):
    """Open the source workbook and locate the ISA 62443-3-2 DRA worksheet.

    Preference order:
    1. A sheet named exactly ``"Worksheet"`` (case-insensitive).
    2. A sheet whose name contains both ``"detail"`` and ``"risk"``.
    3. The first sheet in the workbook.

    Args:
        source_path: Path to the input ``.xlsx`` or ``.xlsm`` workbook.

    Returns:
        A two-tuple of ``(workbook, worksheet)``.
    """
    wb = load_workbook(source_path, data_only=False)
    _validate_workbook_structure(wb, max_sheets=20, max_rows_per_sheet=100000)
    for name in wb.sheetnames:
        if name.lower() == "worksheet":
            return wb, wb[name]
        if "detail" in name.lower() and "risk" in name.lower():
            return wb, wb[name]
    # [IMPROVEMENT] Log a warning when the expected sheet name is not found so
    # operators know the fallback fired and can verify the correct input workbook.
    # Silent fallback could produce misleading enriched output.
    # [SOURCE] Audit finding: Medium (M-5)
    fallback_name = wb.sheetnames[0]
    logger.warning(
        "Source workbook '%s': no sheet named 'Worksheet' or matching 'detail+risk' found."
        " Falling back to first sheet '%s'. Verify this is the correct DRA sheet.",
        source_path,
        fallback_name,
    )
    return wb, wb[fallback_name]


def _parse_source_rows(
    ws,
    technique_catalog: dict[str, dict[str, Any]],
    asset_catalog: dict[str, str],
    mitre_techniques: dict[str, dict[str, Any]],
    mitigation_map: dict[str, list[str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse and enrich every data row from the source ISA 62443-3-2 worksheet.

    Reads data starting at row 4 (rows 1-3 are treated as title/header rows).
    For each row the function:
    - Resolves the asset name from zone-embedded asset IDs or the explicit cell.
    - Enriches the MITRE technique via the builder catalog and the live STIX feed.
    - Derives the adversary model from the free-text threat-source label.
    - Deduplicates rows with the same (risk_id, asset, technique_id) key.

    Args:
        ws: Source openpyxl worksheet.
        technique_catalog: Technique metadata from the builder workbook.
        asset_catalog: Asset ID-to-name mapping from the builder workbook.
        mitre_techniques: Canonical MITRE technique dicts keyed by technique ID.
        mitigation_map: Mitigation names keyed by technique ID.

    Returns:
        A four-tuple of ``(enhanced_rows, mapping_rows, index_rows, change_log)``.
    """

    enhanced_rows: list[dict[str, Any]] = []
    mapping_rows: list[dict[str, Any]] = []
    index_rows: list[dict[str, Any]] = []
    change_log: list[dict[str, Any]] = []

    current_zone = ""
    seen_keys: set[tuple[Any, ...]] = set()

    for row_num in range(4, ws.max_row + 1):
        risk_id = ws[f"A{row_num}"].value
        zone = ws[f"B{row_num}"].value or current_zone
        current_zone = zone or current_zone
        explicit_asset = str(ws[f"C{row_num}"].value or "").strip() or None
        threat_source = ws[f"D{row_num}"].value
        tactic = ws[f"E{row_num}"].value
        technique_id = ws[f"F{row_num}"].value
        technique_name = ws[f"G{row_num}"].value
        technique_description = ws[f"H{row_num}"].value
        vulnerabilities = ws[f"I{row_num}"].value
        consequence_desc = ws[f"J{row_num}"].value
        countermeasures_template = str(ws[f"S{row_num}"].value or "").strip() or None

        if not any([risk_id, zone, threat_source, technique_id, technique_name, vulnerabilities, consequence_desc]):
            continue

        zone_text = str(zone or "")
        asset_ids = _extract_asset_ids(zone_text)
        if explicit_asset:
            asset_names = [explicit_asset]
        elif asset_ids:
            asset_names = [asset_catalog.get(asset_id, asset_id) for asset_id in asset_ids]
        else:
            asset_names = [zone_text.splitlines()[0] if zone_text else "Unmapped Asset"]

        technique_id = str(technique_id).strip() if technique_id else ""
        if technique_id and technique_id in technique_catalog:
            builder_info = technique_catalog[technique_id]
            tactic = tactic or builder_info.get("tactic")
            technique_name = technique_name or builder_info.get("technique_name")
            technique_description = technique_description or builder_info.get("description")
        elif not technique_id:
            scenario_text = f"{vulnerabilities or ''} {consequence_desc or ''}".lower()
            scored = []
            for candidate_id, info in technique_catalog.items():
                overlap = sum(1 for aid in asset_ids if aid in info.get("asset_ids", []))
                keyword_hits = sum(1 for keyword in str(info.get("technique_name", "")).lower().split() if keyword in scenario_text)
                score = overlap * 3 + keyword_hits
                if score > 0:
                    scored.append((score, candidate_id, info))
            if scored:
                scored.sort(reverse=True)
                _, technique_id, info = scored[0]
                tactic = info.get("tactic")
                technique_name = info.get("technique_name")
                technique_description = info.get("description")

        technique_id = technique_id if TECHNIQUE_ID_PATTERN.fullmatch(technique_id or "") else ""
        if technique_id and technique_id in mitre_techniques:
            canonical = mitre_techniques[technique_id]
            tactic = tactic or "; ".join(canonical.get("tactics", []))
            technique_name = technique_name or canonical.get("name")
            technique_description = technique_description or canonical.get("description")

        mitigations = "; ".join(sorted(set(mitigation_map.get(technique_id, [])))) if technique_id else ""
        adversary_name, adversary_type, adversary_description = _normalize_adversary(str(threat_source or "N/A"))
        impact = max([value for value in [ws[f"K{row_num}"].value, ws[f"L{row_num}"].value, ws[f"M{row_num}"].value, ws[f"N{row_num}"].value, ws[f"O{row_num}"].value] if isinstance(value, (int, float))], default="")

        for asset_name in asset_names:
            dedup_key = (risk_id, asset_name, technique_id)
            action = "kept"
            if dedup_key in seen_keys:
                action = "deduplicated"
                change_log.append({
                    "risk_id": risk_id,
                    "source_row": row_num,
                    "action": action,
                    "details": f"Duplicate technique {technique_id or 'UNMAPPED'} for asset {asset_name} skipped.",
                })
                continue
            seen_keys.add(dedup_key)

            enhanced_rows.append(
                {
                    "risk_id": risk_id,
                    "zone": zone_text.splitlines()[0] if zone_text else "",
                    "asset": asset_name,
                    "adversary": adversary_name,
                    "adversary_type": adversary_type,
                    "adversary_description": adversary_description,
                    "technique_id": technique_id,
                    "technique_name": technique_name or "",
                    "technique_description": technique_description or "",
                    "tactics": tactic or "",
                    "mitigations": mitigations,
                    "countermeasures_template": countermeasures_template or "",
                    "impact": impact,
                    "original_threat_source": threat_source or "",
                    "vulnerabilities": vulnerabilities or "",
                    "consequence_description": consequence_desc or "",
                    "source_row": row_num,
                }
            )
            for asset_id in asset_ids or [""]:
                mapping_rows.append(
                    {
                        "risk_id": risk_id,
                        "zone": zone_text.splitlines()[0] if zone_text else "",
                        "asset": asset_name,
                        "mitre_asset_id": asset_id,
                        "technique_id": technique_id,
                        "technique_name": technique_name or "",
                        "dedup_action": action,
                    }
                )
            if technique_id:
                index_rows.append(
                    {
                        "technique_id": technique_id,
                        "technique_name": technique_name or "",
                        "adversary": adversary_name,
                        "mitigations": mitigations,
                    }
                )
            change_log.append(
                {
                    "risk_id": risk_id,
                    "source_row": row_num,
                    "action": "transformed",
                    "details": f"Threat Source replaced with adversary model and MITRE technique {technique_id or 'UNMAPPED'} enriched.",
                }
            )

    return enhanced_rows, mapping_rows, index_rows, change_log


def _extract_worksheet_headers(ws) -> dict[int, str]:
    """Extract the verbatim column headers from the source "Worksheet" sheet.

    Scans the first three rows to identify the header row — defined as the
    row that contains the highest number of non-empty cells.  Returns a
    mapping of 1-based column index to the exact header string so that
    downstream writers can replicate the original column names without any
    normalisation or renaming.

    Args:
        ws: An openpyxl worksheet object (the ``"Worksheet"`` sheet).

    Returns:
        dict mapping 1-based column index to the verbatim header string.
        Columns with empty headers are omitted.
    """
    best_row = 1
    best_count = 0
    for row_num in range(1, 4):
        count = sum(
            1 for col in range(1, ws.max_column + 1) if ws.cell(row_num, col).value
        )
        if count > best_count:
            best_count = count
            best_row = row_num
    return {
        col: str(ws.cell(best_row, col).value).strip()
        for col in range(1, ws.max_column + 1)
        if ws.cell(best_row, col).value is not None
    }


def _write_enhanced_sheet(
    ws,
    source_headers: dict[int, str],
    rows: list[dict[str, Any]],
) -> None:
    """Write the Enhanced Risk Assessment sheet with exact source column headers.

    Fields that originate from the source ``"Worksheet"`` are labelled with
    the verbatim header text extracted from that sheet.  MITRE-enrichment
    fields that have no counterpart in the source use the labels defined in
    :data:`_ENRICHMENT_FIELD_LABELS`.

    Column order follows :data:`_ENHANCED_FIELD_ORDER`.  Column widths are
    auto-sized by sampling up to the first 150 rows.

    Args:
        ws: Target openpyxl worksheet (already reset / empty).
        source_headers: Mapping of 1-based column index to exact header text
            as returned by :func:`_extract_worksheet_headers`.
        rows: Enriched row dicts produced by :func:`_parse_source_rows`
            with sequential ``risk_id`` values already applied.
    """
    output_columns: list[tuple[str, str]] = []
    for field in _ENHANCED_FIELD_ORDER:
        if field in _FIELD_TO_SOURCE_COL:
            src_col = _FIELD_TO_SOURCE_COL[field]
            display = source_headers.get(src_col, field)
        else:
            display = _ENRICHMENT_FIELD_LABELS.get(field, field)
        output_columns.append((display, field))

    ws.append([hdr for hdr, _ in output_columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    # Auto-size columns by sampling the first 150 data rows for performance.
    sample = rows[:150]
    for idx, (hdr, field) in enumerate(output_columns, start=1):
        max_len = max(
            [len(hdr)]
            + [len(str(row.get(field, "") or "")) for row in sample],
            default=len(hdr),
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)

    for row in rows:
        # [IMPROVEMENT] Apply formula-injection guard before writing enriched data cells.
        # STIX descriptions and builder content are external inputs that may contain
        # strings beginning with formula characters.
        # [SOURCE] Audit finding: Critical (C-2)
        ws.append([_safe_cell_value(row.get(field, "")) for _, field in output_columns])


def _reset_sheet(wb, title: str):
    """Delete *title* from *wb* if it exists, then create and return a fresh sheet.

    Args:
        wb: Target openpyxl workbook.
        title: Sheet name to reset.

    Returns:
        The newly created, empty openpyxl worksheet.
    """
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def _write_tabular_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    """Write a list of row dicts to *ws* as a formatted table.

    The first row is written as bold column headers with freeze-panes applied.
    Column widths are auto-sized by sampling up to the first 150 rows.

    Args:
        ws: Target openpyxl worksheet (should be empty / freshly reset).
        headers: Ordered list of column header strings; each is also used as
            the key to look up values in each row dict.
        rows: List of dicts; missing keys default to an empty string.
    """
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        # [IMPROVEMENT] Apply formula-injection guard to every data cell.
        # [SOURCE] Audit finding: Critical (C-2)
        ws.append([_safe_cell_value(row.get(header, "")) for header in headers])

    for idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(ws.cell(r, idx).value or "")) for r in range(2, min(ws.max_row, 150) + 1)])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)


def transform_risk_assessment_workbook(
    source_workbook_path: str,
    template_workbook_path: str,
    mitre_builder_workbook_path: str,
    output_path: str,
) -> dict[str, Any]:
    """Transform an ISA 62443-3-2 Detailed Risk Assessment workbook.

    Orchestrates the full pipeline:
    1. Load the MITRE ATT&CK Builder asset/technique catalog.
    2. Fetch and parse the live MITRE ICS STIX bundle (best-effort).
    3. Parse and enrich every row from the source ``"Worksheet"`` sheet.
    4. Renumber ``risk_id`` sequentially starting from 1.
    5. Extract the verbatim column headers from the source worksheet.
    6. Write an ``Enhanced Risk Assessment`` sheet that preserves exact source
       column header names alongside MITRE-enrichment columns.
    7. Write ``Mapping Reference``, ``MITRE Reference Index``, ``Change Log``,
       and ``Original DRA Snapshot`` sheets into a copy of the template.

    Args:
        source_workbook_path: Path to the input ISA 62443-3-2 workbook
            containing a sheet named ``"Worksheet"``.
        template_workbook_path: Path to the ``.xlsm`` output template that
            carries macros and branding.
        mitre_builder_workbook_path: Path to the MITRE ATT&CK Builder
            ``.xlsm`` containing the ``Config`` and ``ICSMatrix`` sheets.
        output_path: Destination path for the enriched output workbook.

    Returns:
        A summary dict with keys ``output_path``, ``enhanced_rows``,
        ``mapping_rows``, ``techniques_indexed``, and ``change_log_entries``.

    Raises:
        FileNotFoundError: If any of the three input paths does not exist.
        ValueError: Propagated from path-traversal guards in calling code.
    """
    asset_catalog, technique_catalog = _load_builder_catalog(mitre_builder_workbook_path)
    mitre_techniques, mitigation_map = _load_mitre_reference()
    source_wb, source_ws = _find_source_sheet(source_workbook_path)

    # Extract verbatim column headers from the source sheet BEFORE parsing rows
    # so that the exact header text is available when building the output sheet.
    source_headers = _extract_worksheet_headers(source_ws)

    enhanced_rows, mapping_rows, index_rows, change_log = _parse_source_rows(
        source_ws,
        technique_catalog=technique_catalog,
        asset_catalog=asset_catalog,
        mitre_techniques=mitre_techniques,
        mitigation_map=mitigation_map,
    )

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Renumber risk_id sequentially (1, 2, 3 …) in the output
    for idx, row in enumerate(enhanced_rows, start=1):
        row["risk_id"] = idx

    template_wb = load_workbook(template_workbook_path, keep_vba=True, data_only=False)
    _validate_workbook_structure(template_wb, max_sheets=10, max_rows_per_sheet=10000)

    enhanced_ws = _reset_sheet(template_wb, "Enhanced Risk Assessment")
    _write_enhanced_sheet(enhanced_ws, source_headers, enhanced_rows)

    mapping_ws = _reset_sheet(template_wb, "Mapping Reference")
    _write_tabular_sheet(
        mapping_ws,
        ["risk_id", "zone", "asset", "mitre_asset_id", "technique_id", "technique_name", "dedup_action"],
        mapping_rows,
    )

    ref_ws = _reset_sheet(template_wb, "MITRE Reference Index")
    unique_index_rows = []
    seen_index = set()
    for row in index_rows:
        key = (row["technique_id"], row["adversary"], row["mitigations"])
        if key not in seen_index:
            seen_index.add(key)
            unique_index_rows.append(row)
    _write_tabular_sheet(ref_ws, ["technique_id", "technique_name", "adversary", "mitigations"], unique_index_rows)

    log_ws = _reset_sheet(template_wb, "Change Log")
    _write_tabular_sheet(log_ws, ["risk_id", "source_row", "action", "details"], change_log)

    snapshot_title = "Original DRA Snapshot"
    if snapshot_title in template_wb.sheetnames:
        del template_wb[snapshot_title]
    snapshot_ws = template_wb.create_sheet(snapshot_title)
    for row in source_ws.iter_rows(values_only=False):
        snapshot_ws.append([cell.value for cell in row])

    template_wb.save(out_path)
    template_wb.close()
    source_wb.close()

    return {
        "output_path": str(out_path),
        "enhanced_rows": len(enhanced_rows),
        "mapping_rows": len(mapping_rows),
        "techniques_indexed": len(unique_index_rows),
        "change_log_entries": len(change_log),
    }
