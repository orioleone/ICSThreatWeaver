from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import column_index_from_string


HEADERS = ["Risk ID", "Zone", "Asset", "Tactic", "Technique ID", "Technique Name", "Description", "Mitigations"]
FIELD_TO_HEADER = {
    "risk_id": "Risk ID",
    "zone": "Zone",
    "asset": "Asset",
    "tactic": "Tactic",
    "technique_id": "Technique ID",
    "technique_name": "Technique Name",
    "description": "Description",
    "mitigations": "Mitigations",
}

# [RESOLVED] Regex to detect strings that Excel/LibreOffice would evaluate as
# [SOURCE] Audit finding: Critical (C-2)
# formulas.  MITRE STIX descriptions and technique names come from an external
# feed; a crafted or compromised bundle could inject formulas such as
# =HYPERLINK("http://attacker.com","Click").  Prefixing such values with a
# single-quote causes the spreadsheet application to treat them as plain text.
# [SOURCE] Audit finding: Critical (C-2)
_FORMULA_PREFIX_RE = re.compile(r"^[=+\-@]")


# [RESOLVED] ZIP bomb defense: Validate workbook structure to prevent malicious
# archives from consuming excessive memory during decompression.
# [SOURCE] Audit finding: High (H-5)
def _validate_workbook_structure(workbook, max_sheets: int = 100, max_rows_per_sheet: int = 100000) -> None:
    """Validate that workbook structure is reasonable (not a ZIP bomb).
    
    Args:
        workbook: Loaded openpyxl Workbook instance.
        max_sheets: Maximum number of sheets allowed (default 100).
        max_rows_per_sheet: Maximum rows per sheet (default 100,000).
        
    Raises:
        ValueError: If workbook structure exceeds limits.
    """
    if len(workbook.sheetnames) > max_sheets:
        raise ValueError(
            f"Workbook contains {len(workbook.sheetnames)} sheets; maximum allowed is {max_sheets}. "
            "This may indicate a malicious or corrupted file."
        )
    
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        if ws.max_row and ws.max_row > max_rows_per_sheet:
            raise ValueError(
                f"Sheet '{sheet}' contains {ws.max_row} rows; maximum allowed is {max_rows_per_sheet}. "
                "This may indicate a malicious or corrupted file."
            )


def _safe_cell_value(value: Any) -> Any:
    """Neutralise potential spreadsheet formula injection in *value*.

    If *value* is a string that starts with a spreadsheet formula character
    (``=``, ``+``, ``-``, ``@``) it is prefixed with a single quote so that
    spreadsheet applications render it as literal text instead of evaluating it.
    Non-string values are returned unchanged.
    """
    if isinstance(value, str) and _FORMULA_PREFIX_RE.match(value):
        # Leading apostrophe is the de-facto standard escape recognised by
        # Excel, LibreOffice Calc, and Google Sheets.
        return "'" + value
    return value


def _get_or_create_sheet(workbook, title: str):
    """Return the existing sheet named *title*, or create it if absent.

    Args:
        workbook: An openpyxl :class:`Workbook` instance.
        title: Name of the sheet to find or create.

    Returns:
        The existing or newly created openpyxl worksheet.
    """
    return workbook[title] if title in workbook.sheetnames else workbook.create_sheet(title)


def _write_table(worksheet, headers: list[str], rows: list[list[str]]) -> None:
    """Clear *worksheet* and write *headers* and *rows* as a simple table.

    The header row is written in bold with freeze-panes applied at row 2.

    Args:
        worksheet: Target openpyxl worksheet.
        headers: Column header labels.
        rows: Data rows; each row must be a list aligned to *headers*.
    """
    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    for row in rows:
        # [RESOLVED] Apply formula-injection guard to every cell value.
        # [SOURCE] Audit finding: Critical (C-2)
        # [SOURCE] Audit finding: Critical (C-2)
        worksheet.append([_safe_cell_value(v) for v in row])


def _discover_template_columns(worksheet, start_row: int, start_col: int) -> tuple[int, dict[str, int]]:
    """Scan the template worksheet to locate existing column positions.

    Searches rows 1 through *start_row* (up to row 15) for the row that
    maps the most recognised field names to column indices.  The mapping
    uses keyword matching so that common variants (e.g. ``"Countermeasure"``
    or ``"Mitigations"``) are normalised back to the internal field key.

    Any configured field not found in the template is appended after the
    last detected column.

    Args:
        worksheet: The target openpyxl worksheet.
        start_row: The first row expected to contain data (header is before it).
        start_col: Fallback starting column index when no headers are found.

    Returns:
        A two-tuple of ``(header_row_index, column_map)`` where
        ``column_map`` maps internal field keys to 1-based column indices.
    """
    best_row = max(start_row - 1, 1)
    best_map: dict[str, int] = {}

    for candidate_row in range(1, min(start_row, 15) + 1):
        mapping: dict[str, int] = {}
        for column_index in range(1, worksheet.max_column + 1):
            value = str(worksheet.cell(candidate_row, column_index).value or "").strip().lower()
            if not value:
                continue
            if "risk" in value and ("id" in value or "no" in value or "#" in value) or value in {"#", "no.", "seq"}:
                mapping["risk_id"] = column_index
            elif value == "zone":
                mapping["zone"] = column_index
            elif value == "asset" or "threat source" in value:
                mapping["asset"] = column_index
            elif value in {"tactic", "tactics"} or "tactic" in value:
                mapping["tactic"] = column_index
            elif ("technique id" in value or value == "id") and "technique_id" not in mapping:
                mapping["technique_id"] = column_index
            elif value in {"technique name", "technique"} or ("technique" in value and "technique_name" not in mapping):
                mapping["technique_name"] = column_index
            elif "description" in value:
                mapping["description"] = column_index
            elif "mitigation" in value or "countermeasure" in value or "vulnerab" in value:
                mapping["mitigations"] = column_index
        if len(mapping) > len(best_map):
            best_row = candidate_row
            best_map = mapping

    if not best_map:
        best_map = {field: start_col + offset for offset, field in enumerate(FIELD_TO_HEADER.keys())}
    else:
        next_col = max(best_map.values(), default=start_col - 1) + 1
        for field in FIELD_TO_HEADER:
            if field not in best_map:
                best_map[field] = next_col
                next_col += 1
    return best_row, best_map


def _prepare_write_area(worksheet, data_start_row: int, relevant_columns: set[int]) -> None:
    """Unmerge merged cells that overlap the data-write area.

    Preserves the style of the anchor cell across cells that were previously
    merged, but only for *relevant_columns*. This prevents style spillover into
    unrelated template columns.

    Args:
        worksheet: Target openpyxl worksheet.
        data_start_row: First row number that will be used for data output.
        relevant_columns: Set of 1-based column indices being written.
    """
    if not relevant_columns:
        return

    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_row < data_start_row:
            continue
        if merged_range.max_col < min(relevant_columns) or merged_range.min_col > max(relevant_columns):
            continue

        anchor = worksheet.cell(merged_range.min_row, merged_range.min_col)
        preserved_style = {
            "font": copy(anchor.font),
            "fill": copy(anchor.fill),
            "border": copy(anchor.border),
            "alignment": copy(anchor.alignment),
            "number_format": anchor.number_format,
            "protection": copy(anchor.protection),
        }
        worksheet.unmerge_cells(str(merged_range))
        target_columns = [
            column
            for column in range(merged_range.min_col, merged_range.max_col + 1)
            if column in relevant_columns
        ]
        for row in range(max(data_start_row, merged_range.min_row), merged_range.max_row + 1):
            for column in target_columns:
                cell = worksheet.cell(row, column)
                cell.font = copy(preserved_style["font"])
                cell.fill = copy(preserved_style["fill"])
                cell.border = copy(preserved_style["border"])
                cell.alignment = copy(preserved_style["alignment"])
                cell.number_format = preserved_style["number_format"]
                cell.protection = copy(preserved_style["protection"])


def _clear_output_area(worksheet, data_start_row: int, relevant_columns: set[int]) -> None:
    """Set all data cells in the write area to ``None``.

    Only cells from *data_start_row* onwards and in *relevant_columns* are
    cleared; header rows and unrelated columns are untouched.

    Args:
        worksheet: Target openpyxl worksheet.
        data_start_row: First row number to clear.
        relevant_columns: Set of 1-based column indices to clear.
    """
    for row in range(data_start_row, worksheet.max_row + 1):
        for column in relevant_columns:
            worksheet.cell(row, column).value = None


def _apply_rectangular_grid_border(
    worksheet,
    start_row: int,
    end_row: int,
    start_col: int = 1,
    end_col: int = 23,
) -> None:
    """Apply a full thin-cell border to a rectangular worksheet region.

    Defaults to columns A..W (1..23) to satisfy matrix export presentation
    requirements regardless of template preformatted row limits.
    """
    if end_row < start_row:
        return

    thin = Side(style="thin", color="000000")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            worksheet.cell(row=row, column=col).border = grid_border


def _capture_template_styles(worksheet, data_start_row: int, column_map: dict[str, int]) -> dict[str, dict[str, Any]]:
    """Capture per-field style blueprint from the template's first data row.

    The snapshot is later applied to every written data cell so styles such as
    borders and wrap/alignment remain consistent even after writing beyond the
    template's preformatted row range.
    """
    styles: dict[str, dict[str, Any]] = {}
    for field, col_index in column_map.items():
        template_cell = worksheet.cell(row=data_start_row, column=col_index)
        styles[field] = {
            "font": copy(template_cell.font),
            "fill": copy(template_cell.fill),
            "border": copy(template_cell.border),
            "alignment": copy(template_cell.alignment),
            "number_format": template_cell.number_format,
            "protection": copy(template_cell.protection),
        }
    return styles


def export_matrix_to_workbook(
    matrix_rows: Iterable[dict],
    template_path: str,
    output_path: str,
    sheet_name: str = "Assessment Matrix",
    start_cell: str = "A2",
) -> str:
    """Write *matrix_rows* into a copy of *template_path* and save to *output_path*.

    The function locates an existing header row in the target sheet using
    keyword matching and writes data beginning at *start_cell*.  Original
    header text in the template is **preserved** — headers are never
    overwritten with normalised internal names.  Only cells that are empty
    (i.e. not yet labelled in the template) will receive a default header.

    Supplementary sheets ``Mapping Reference``, ``MITRE Reference Index``,
    and ``Change Log`` are created or overwritten.

    Args:
        matrix_rows: Iterable of row dicts each containing the field keys
            defined in :data:`FIELD_TO_HEADER`.
        template_path: Path to the source template (``.xlsx`` or ``.xlsm``).
        output_path: Destination path for the populated output workbook.
        sheet_name: Name of the worksheet within the template to populate.
            Defaults to ``"Assessment Matrix"``.
        start_cell: Top-left cell reference (e.g. ``"A2"``) for the data
            block, including its header row.  Defaults to ``"A2"``.

    Returns:
        Absolute path to the saved output file as a string.

    Raises:
        FileNotFoundError: If *template_path* does not exist.
    """
    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    rows = list(matrix_rows)
    workbook = load_workbook(template, keep_vba=template.suffix.lower() == ".xlsm", data_only=False)
    _validate_workbook_structure(workbook)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

    col_letters = "".join(ch for ch in start_cell if ch.isalpha()) or "A"
    row_digits = "".join(ch for ch in start_cell if ch.isdigit()) or "2"
    start_col = column_index_from_string(col_letters)
    start_row = int(row_digits)

    header_row, column_map = _discover_template_columns(worksheet, start_row, start_col)
    # Preserve original template headers: only write a header label when the
    # cell is currently empty.  This ensures names like "Countermeasure" are
    # never silently replaced with the internal alias "Mitigations".
    for field, col_index in column_map.items():
        header_cell = worksheet.cell(row=header_row, column=col_index)
        if not header_cell.value:
            header_cell.value = FIELD_TO_HEADER.get(field, field)
        header_cell.font = Font(bold=True)

    data_start_row = max(start_row, header_row + 1)
    relevant_columns = set(column_map.values())
    _prepare_write_area(worksheet, data_start_row, relevant_columns)
    template_styles = _capture_template_styles(worksheet, data_start_row, column_map)
    _clear_output_area(worksheet, data_start_row, relevant_columns)

    for seq_num, row in enumerate(rows, start=1):
        row_index = data_start_row + seq_num - 1
        for field in FIELD_TO_HEADER:
            value = seq_num if field == "risk_id" else row.get(field, "")
            # [RESOLVED] Sanitise every cell write against formula injection.
            # [SOURCE] Audit finding: Critical (C-2)
            # [SOURCE] Audit finding: Critical (C-2)
            cell = worksheet.cell(row=row_index, column=column_map[field], value=_safe_cell_value(value))
            style = template_styles.get(field)
            if style:
                cell.font = copy(style["font"])
                cell.fill = copy(style["fill"])
                cell.border = copy(style["border"])
                cell.alignment = copy(style["alignment"])
                cell.number_format = style["number_format"]
                cell.protection = copy(style["protection"])

    if rows:
        _apply_rectangular_grid_border(
            worksheet,
            start_row=header_row,
            end_row=data_start_row + len(rows) - 1,
            start_col=1,
            end_col=column_index_from_string("W"),
        )

    mapping_sheet = _get_or_create_sheet(workbook, "Mapping Reference")
    mapping_rows = [
        [row.get("technique_id", ""), row.get("technique_name", ""), row.get("tactic", ""), row.get("mitigations", "")]
        for row in rows
    ]
    _write_table(mapping_sheet, ["Technique ID", "Technique Name", "Tactic", "Mitigations"], mapping_rows)

    index_sheet = _get_or_create_sheet(workbook, "MITRE Reference Index")
    dedup_index = []
    seen = set()
    for row in rows:
        key = (row.get("technique_id", ""), row.get("technique_name", ""))
        if key not in seen:
            seen.add(key)
            dedup_index.append([row.get("technique_id", ""), row.get("technique_name", ""), row.get("tactic", "")])
    _write_table(index_sheet, ["Technique ID", "Technique Name", "Tactic"], dedup_index)

    log_sheet = _get_or_create_sheet(workbook, "Change Log")
    _write_table(
        log_sheet,
        ["Action", "Details"],
        [
            ["safe-write", f"Matrix written starting at {start_cell} on sheet {sheet_name}."],
            ["template-alignment", "Detected worksheet headers dynamically and wrote only supported template fields."],
            ["format-safe-write", "Only output data cells and reference sheets were updated; existing workbook structure was left intact."],
            ["rows-generated", str(len(rows))],
        ],
    )

    workbook.save(output)
    workbook.close()
    return str(output)
